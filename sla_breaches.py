import requests
import openpyxl
from time import sleep
import sys
from time import strftime, localtime
from email_utility import Email_Utility

def main():
    
    while True:
        time = strftime("%Y/%m/%d %H:%M:%S", localtime())
        
        curr_excel = 'G:\everyone\~CFM\Process Management\TableauFiles\CCC SIIAM SLA Breaches V1.xlsx'
        new_excel_link = 'http://ccc-tnovm.wg.dir.telstra.com:5010/webform_csv/sla_breach_register.xlsx'
        
        r = requests.get(new_excel_link, allow_redirects=True)
        open('CCC SIIAM SLA Breaches Updated.xlsx', 'wb').write(r.content)
        
        new_excel = 'CCC SIIAM SLA Breaches Updated.xlsx'
        
        # Columns
        submit_date_col = 1
        submitter_email_col = 2
        submitter_id_col = 3
        siiam_case_id_col = 4
        closed_date_col = 5
        high_level_group_col = 6
        exception_col = 7
        field_col = 8
        network_col = 9
        description_col = 10
        conen_col = 11
        
        # Current excel file
        try:
            curr_book = openpyxl.load_workbook(curr_excel)
        except FileNotFoundError as e:
            print(f'Excel file "{curr_excel}" does not exist!')
            # Send email to admin to advise
            email = Email_Utility(['franz.weishaupl@team.telstra.com'], [''], ['peter.lynch@team.telstra.com', 'kamran.khan@team.telstra.com'],
                                  'CCC SIIAM SLA Breaches Script - Error',
                                  f'''This is an automated message to advise that "CCC SIIAM SLA Breaches Script" has encountered an exception.

The exception is as below:
{e}

Please restart the script on robot workstation with IP ADDRESS 172.115.25.35

Regards,
Franz''', f'', f'')
            email.Send()
        curr_sheet = curr_book['query'] # Get 'query' tab
#        curr_max_row = curr_sheet.max_row # No of written Rows in sheet
    #    print(f'curr_max_row: {curr_max_row}')
        
        # New excel file
        new_book = openpyxl.load_workbook(new_excel)
        new_sheet = new_book['query'] # Get 'query' tab
        new_max_row = new_sheet.max_row # No of written Rows in sheet
    #    print(f'new_max_row: {new_max_row}')
    
        print(f'\n\nTimestamp: {time}\nUpdating excel file...')
        for record in range(1,new_max_row+1,1):
            # Get latest entry in new sheet
            submit_date_new = new_sheet.cell(row=record,column=submit_date_col)
            submitter_email_new = new_sheet.cell(row=record,column=submitter_email_col)
            submitter_id_new = new_sheet.cell(row=record,column=submitter_id_col)
            siiam_case_id_new = new_sheet.cell(row=record,column=siiam_case_id_col)
            closed_date_new = new_sheet.cell(row=record,column=closed_date_col)
            high_level_group_new = new_sheet.cell(row=record,column=high_level_group_col)
            exception_new = new_sheet.cell(row=record,column=exception_col)
            field_new = new_sheet.cell(row=record,column=field_col)
            network_new = new_sheet.cell(row=record,column=network_col)
            description_new = new_sheet.cell(row=record,column=description_col)
            conen_new = new_sheet.cell(row=record,column=conen_col)

            # Append to current excel file
            curr_sheet.cell(row=record, column=submit_date_col).value = submit_date_new.value
            curr_sheet.cell(row=record, column=submitter_email_col).value = submitter_email_new.value
            curr_sheet.cell(row=record, column=submitter_id_col).value = submitter_id_new.value
            curr_sheet.cell(row=record, column=siiam_case_id_col).value = siiam_case_id_new.value
            curr_sheet.cell(row=record, column=closed_date_col).value = closed_date_new.value
            curr_sheet.cell(row=record, column=high_level_group_col).value = high_level_group_new.value
            curr_sheet.cell(row=record, column=exception_col).value = exception_new.value
            curr_sheet.cell(row=record, column=field_col).value = field_new.value
            curr_sheet.cell(row=record, column=network_col).value = network_new.value
            curr_sheet.cell(row=record, column=description_col).value = description_new.value
            curr_sheet.cell(row=record, column=conen_col).value = conen_new.value
            while True:
                try:
                    curr_book.save('G:\everyone\~CFM\Process Management\TableauFiles\CCC SIIAM SLA Breaches V1.xlsx')
                    break
                except PermissionError as e:
                    print(f'\n\nCannot update "{curr_excel}" as it is currently open.')
                    for remaining in range(60, 0, -1):
                        sys.stdout.write("\r")
                        sys.stdout.write("Retrying in {:2d} seconds...".format(remaining)) 
                        sys.stdout.flush()
                        sleep(1)
        print(f'Completed updating: {curr_excel}')
        print(f'Spreadsheet has {record} records.\n')
            
        for remaining in range(3600, 0, -1):
            sys.stdout.write("\r")
            sys.stdout.write("{:2d} seconds remaining until next update...".format(remaining)) 
            sys.stdout.flush()
            sleep(1)

if __name__ == '__main__':
    main()
    
